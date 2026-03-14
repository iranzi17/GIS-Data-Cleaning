from __future__ import annotations

import argparse
import re
import warnings
from pathlib import Path

import geopandas as gpd
import pyogrio
from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore",
    message="Unknown extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl",
)


START_DEVICE = "Voltage Transformer"
VT_ID_FIELD = "VoltageTransfomer_ID"
SUBSTATION_FIELD = "Substation_ID"
TYPE_FIELD = "TypeofVoltageTransformer"


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def first_non_blank(*values: object) -> object:
    for value in values:
        if not is_blank(value):
            return value
    return None


def coerce_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None

    text = str(value).strip()
    if not text:
        return None

    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else None


def iter_voltage_transformer_blocks(workbook_path: Path) -> list[dict[str, dict[str, object]]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    blocks: list[dict[str, dict[str, object]]] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        index = 0

        while index < len(rows):
            row = rows[index]
            cell_a = row[0] if len(row) > 0 else None
            cell_b = row[1] if len(row) > 1 else None

            if normalize(cell_a) != normalize(START_DEVICE) or normalize(cell_b) != normalize(VT_ID_FIELD):
                index += 1
                continue

            block: dict[str, dict[str, object]] = {}
            cursor = index

            while cursor < len(rows):
                current = rows[cursor]
                cells = list(current[:5]) + [None] * max(0, 5 - len(current[:5]))

                if cursor > index and all(is_blank(value) for value in cells[:5]):
                    break

                field_name = cells[1]
                if not is_blank(field_name):
                    block[str(field_name).strip()] = {
                        "value": first_non_blank(cells[4], cells[3]),
                        "domain": cells[3],
                        "domain_code": cells[4],
                        "row_number": cursor + 1,
                        "sheet_name": worksheet.title,
                    }

                cursor += 1

            if block:
                blocks.append(block)

            index = cursor

    return blocks


def build_type_label_map(workbook_paths: list[Path]) -> dict[str, int]:
    label_map: dict[str, int] = {}

    for workbook_path in workbook_paths:
        for block in iter_voltage_transformer_blocks(workbook_path):
            type_row = block.get(TYPE_FIELD)
            if not type_row:
                continue

            label = normalize(type_row.get("domain"))
            code = coerce_int(type_row.get("domain_code"))
            if not label or code is None:
                continue

            existing = label_map.get(label)
            if existing is not None and existing != code:
                raise ValueError(
                    f"Conflicting type code mapping for '{type_row.get('domain')}' in {workbook_path.name}: "
                    f"{existing} vs {code}"
                )
            label_map[label] = code

    return label_map


def resolve_type_code(type_row: dict[str, object], label_map: dict[str, int]) -> int | None:
    direct_code = coerce_int(type_row.get("domain_code"))
    if direct_code is not None:
        return direct_code

    for candidate in (type_row.get("domain"), type_row.get("domain_code"), type_row.get("value")):
        normalized = normalize(candidate)
        if normalized in label_map:
            return label_map[normalized]

    return None


def collect_workbook_records(workbook_paths: list[Path], label_map: dict[str, int]) -> dict[tuple[str, str], int]:
    records: dict[tuple[str, str], int] = {}

    for workbook_path in workbook_paths:
        for block in iter_voltage_transformer_blocks(workbook_path):
            substation_row = block.get(SUBSTATION_FIELD)
            vt_id_row = block.get(VT_ID_FIELD)
            type_row = block.get(TYPE_FIELD)

            if not substation_row or not vt_id_row or not type_row:
                continue

            substation_id = first_non_blank(substation_row.get("value"), substation_row.get("domain"))
            vt_id = first_non_blank(vt_id_row.get("value"), vt_id_row.get("domain"))
            type_code = resolve_type_code(type_row, label_map)

            if is_blank(substation_id) or is_blank(vt_id) or type_code is None:
                continue

            key = (normalize(substation_id), normalize(vt_id))
            existing = records.get(key)
            if existing is not None and existing != type_code:
                raise ValueError(
                    f"Conflicting workbook codes for {substation_id} / {vt_id}: {existing} vs {type_code}"
                )
            records[key] = type_code

    return records


def matching_workbooks(gdb_path: Path, workbook_paths: list[Path]) -> list[Path]:
    gdb_key = normalize(gdb_path.stem)
    return [path for path in workbook_paths if normalize(path.stem).startswith(gdb_key)]


def update_voltage_transformer_layer(gdb_path: Path, workbook_record_map: dict[tuple[str, str], int]) -> list[tuple[str, str, int]]:
    dataframe = gpd.read_file(gdb_path, layer="voltage_transformer")
    updates: list[tuple[str, str, int]] = []

    for index, row in dataframe[dataframe[TYPE_FIELD].isna()].iterrows():
        key = (normalize(row[SUBSTATION_FIELD]), normalize(row[VT_ID_FIELD]))
        code = workbook_record_map.get(key)
        if code is None:
            continue

        dataframe.at[index, TYPE_FIELD] = code
        updates.append((str(row[SUBSTATION_FIELD]), str(row[VT_ID_FIELD]), code))

    if updates:
        pyogrio.write_dataframe(dataframe, gdb_path, layer="voltage_transformer", driver="OpenFileGDB")

    return updates


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill missing TypeofVoltageTransformer values in Conductors geodatabases from New Data workbooks."
    )
    parser.add_argument(
        "--conductors-dir",
        default="Conductors",
        type=Path,
        help="Directory containing target .gdb folders.",
    )
    parser.add_argument(
        "--new-data-dir",
        default="New Data",
        type=Path,
        help="Directory containing the source Excel workbooks.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    conductors_dir = args.conductors_dir
    new_data_dir = args.new_data_dir

    workbook_paths = sorted(
        path for path in new_data_dir.glob("*DATA.xlsx") if not path.name.startswith("~$")
    )
    if not workbook_paths:
        raise FileNotFoundError(f"No workbook files found in {new_data_dir}")

    label_map = build_type_label_map(workbook_paths)
    if not label_map:
        raise ValueError("No voltage transformer type label mappings were discovered in the workbooks.")

    print("Discovered type label mapping:")
    for label, code in sorted(label_map.items()):
        print(f"  {label} -> {code}")

    total_updates = 0
    for gdb_path in sorted(conductors_dir.glob("*.gdb")):
        related_workbooks = matching_workbooks(gdb_path, workbook_paths)
        if not related_workbooks:
            print(f"{gdb_path.name}: no matching workbook found")
            continue

        workbook_record_map = collect_workbook_records(related_workbooks, label_map)
        updates = update_voltage_transformer_layer(gdb_path, workbook_record_map)

        if not updates:
            print(f"{gdb_path.name}: no missing values updated")
            continue

        total_updates += len(updates)
        print(f"{gdb_path.name}: updated {len(updates)} row(s)")
        for substation_id, vt_id, code in updates:
            print(f"  {substation_id} / {vt_id} -> {code}")

    print(f"Total updated rows: {total_updates}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
